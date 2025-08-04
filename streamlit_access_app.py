import streamlit as st
import pandas as pd
import json
import os
import random
from openpyxl import load_workbook, Workbook
from datetime import datetime
from io import BytesIO

# Constants
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_CONFIG = os.path.join(BASE_DIR, 'active_config.json')
CODE_FILE = os.path.join(BASE_DIR, 'supervisor_codes.json')
EXCEL_OUTPUT = os.path.join(BASE_DIR, 'access_review_log.xlsx')

# CSS Styling
st.set_page_config(page_title="Access Review Portal", layout="centered")
st.markdown(
    """
    <style>
    /* Page background */
    .stApp {
        background-color: #e0f7ff;
    }
    /* Main content container */
    .block-container {
        background-color: #ffffff !important;
        border-radius: 15px;
        padding: 2rem !important;
    }
    /* Tabs container */
    .stTabs, .stTabs [role="tablist"] > button {
        background-color: #ffffff !important;
        border-radius: 15px;
        margin: 0 0.25rem;
    }
    /* Dark blue buttons */
    .stButton > button {
        background-color: #004085;
        color: #ffffff;
        border-radius: 8px;
    }
    .stButton > button:hover {
        background-color: #003366;
    }
    .stButton > button:focus {
        outline: 2px solid #003366;
    }
    </style>
    """, unsafe_allow_html=True
)

# Utilities
def generate_unique_code(existing):
    while True:
        code = str(random.randint(1000, 9999))
        if code not in existing:
            return code


def load_active_csv_path():
    if not os.path.exists(DATA_CONFIG):
        st.error("No active config found. Upload a CSV first.")
        st.stop()
    with open(DATA_CONFIG, 'r') as f:
        config = json.load(f)
    return config.get('active_csv')


def load_dataframe():
    csv_path = load_active_csv_path()
    return pd.read_csv(csv_path, encoding='ISO-8859-1')


def load_or_create_codes(df):
    # 1) Clean up the column names
    df.columns = df.columns.str.replace(r'^\ufeff', '', regex=True).str.strip()
    # 2) Find the real “Supervisor” column (case-insensitive)
    sup_col = next((c for c in df.columns if c.lower() == 'supervisor'), None)
    if sup_col is None:
        st.error("Missing ‘Supervisor’ column in your data after normalization.")
        st.stop()

    # 3) Extract unique supervisors
    supervisors = sorted(df[sup_col].dropna().unique())

    # 4) Load or build the code map
    code_map = {}
    if os.path.exists(CODE_FILE):
        with open(CODE_FILE, 'r') as f:
            code_map = json.load(f)
    for sup in supervisors:
        if sup not in code_map:
            code_map[sup] = generate_unique_code(code_map.values())

    with open(CODE_FILE, 'w') as f:
        json.dump(code_map, f, indent=2)

    return code_map



def find_supervisor_by_code(code):
    if not os.path.exists(CODE_FILE):
        return None
    with open(CODE_FILE, 'r') as f:
        code_map = json.load(f)
    return next((sup for sup, sup_code in code_map.items() if sup_code == code), None)


def log_actions(supervisor, approved, removed, df):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sup_df = df[df['Supervisor'] == supervisor][['User ID', 'User Name', 'Role', 'Role Name']].drop_duplicates()
    records = []
    for action, lst in [('Approved', approved), ('Removed', removed)]:
        for entry in lst:
            uid, _ = entry.split(' - ', 1)
            match = sup_df[sup_df['User ID'] == uid]
            if not match.empty:
                row = match.iloc[0].to_dict()
                row.update({'Supervisor': supervisor, 'Action': action, 'Timestamp': timestamp})
                records.append(row)
    if not records:
        return
    if not os.path.exists(EXCEL_OUTPUT):
        wb = Workbook()
        ws = wb.active
        ws.append(['User ID', 'User Name', 'Role', 'Role Name', 'Supervisor', 'Action', 'Timestamp'])
        for rec in records:
            ws.append([rec['User ID'], rec['User Name'], rec['Role'], rec['Role Name'], rec['Supervisor'], rec['Action'], rec['Timestamp']])
        wb.save(EXCEL_OUTPUT)
    else:
        wb = load_workbook(EXCEL_OUTPUT)
        ws = wb.active
        for rec in records:
            ws.append([rec['User ID'], rec['User Name'], rec['Role'], rec['Role Name'], rec['Supervisor'], rec['Action'], rec['Timestamp']])
        wb.save(EXCEL_OUTPUT)

# Session state defaults
for key, default in [('supervisor', None), ('review_started', False), ('review_complete', False), ('approved', []), ('removed', [])]:
    if key not in st.session_state:
        st.session_state[key] = default

# Streamlit UI
st.title("Access Review Portal")
tabs = st.tabs(["Supervisor Review", "Admin Panel"])

# Supervisor Review Tab
with tabs[0]:
    st.header("Supervisor Access Review")
    if not st.session_state.review_started:
        code = st.text_input("Enter your 4-digit access code", max_chars=4)
        if st.button("Start Review") and code:
            sup = find_supervisor_by_code(code)
            if sup:
                st.session_state.supervisor = sup
                st.session_state.review_started = True
                st.rerun()
            else:
                st.error("Invalid access code.")
    elif st.session_state.review_started and not st.session_state.review_complete:
        sup = st.session_state.supervisor
        df = load_dataframe()
        emps = df[df['Supervisor'] == sup][['User ID','User Name','Role','Role Name']].drop_duplicates()
        st.success(f"Welcome, {sup}. Review your team’s access:")
        with st.form("access_form"):
            approve_all = st.checkbox("Approve all", key='approve_all')
            approve, remove = [], []
            for i, row in emps.iterrows():
                label = f"{row['User ID']} - {row['User Name']}"
                c1, c2 = st.columns(2)
                with c1:
                    if st.checkbox("Approve", key=f"a_{i}", value=approve_all): approve.append(label)
                with c2:
                    if st.checkbox("Remove", key=f"r_{i}"): remove.append(label)
                st.markdown(f"**{label}** | {row['Role']} - {row['Role Name']}")
            if st.form_submit_button("Submit Review"):
                log_actions(sup, approve, remove, df)
                st.session_state.approved, st.session_state.removed = approve, remove
                st.session_state.review_complete = True
                st.rerun()
    else:
        st.success("✅ Review submitted!")
        st.write(f"**Supervisor:** {st.session_state.supervisor}")
        if st.session_state.approved:
            st.markdown("**Approved:**")
            st.write(st.session_state.approved)
        if st.session_state.removed:
            st.markdown("**Removed:**")
            st.write(st.session_state.removed)
        if st.button("Start New Review"):
            for k in ['review_started','review_complete','approved','removed']:
                st.session_state[k] = False if isinstance(st.session_state[k], bool) else []
            st.rerun()

# Admin Panel Tab
with tabs[1]:
    st.header("Admin Panel")
    # Initialize admin passcode state
    if 'admin_verified' not in st.session_state:
        st.session_state.admin_verified = False

    # Step 1: Passcode entry
    if not st.session_state.admin_verified:
        pwd = st.text_input("Admin passcode", type="password", key='admin_pass')
        if st.button("Unlock Admin Panel", key='unlock_button'):
            if pwd == "1234":
                st.session_state.admin_verified = True
                st.rerun()
            else:
                st.error("Incorrect passcode.")
    # Step 2: CSV upload & code generation
    else:
        st.subheader("Upload Employee CSVs")
        with st.form("upload_form"):
            uploaded_files = st.file_uploader(
                "Choose CSV files", type=['csv'], accept_multiple_files=True, key='csv_upload'
            )
            upload_submit = st.form_submit_button("Upload Files")

        if upload_submit:
            if not uploaded_files:
                st.warning("Please select at least one CSV file.")
            else:
                all_dfs = []
                file_data = []
                expected_cols = None

                # Read and normalize each CSV
                for idx, uploaded in enumerate(uploaded_files):
                    data_bytes = uploaded.read()
                    if idx == 0:
                        # first file: read normally
                        df_part = pd.read_csv(
                            BytesIO(data_bytes),
                            encoding='ISO-8859-1',
                            engine='python',
                            on_bad_lines='skip'
                        )
                        # capture and normalize column names
                        df_part.columns = (
                            df_part.columns
                                .str.replace(r'^﻿', '', regex=True)
                                .str.strip()
                        )
                        expected_cols = df_part.columns.tolist()
                    else:
                        # subsequent files: skip their header row, enforce columns
                        df_part = pd.read_csv(
                            BytesIO(data_bytes),
                            encoding='ISO-8859-1',
                            engine='python',
                            on_bad_lines='skip',
                            header=None,
                            names=expected_cols,
                            skiprows=1
                        )
                    all_dfs.append(df_part)
                    file_data.append((uploaded.name, data_bytes))

                if not all_dfs:
                    st.error("No valid CSV files could be parsed.")
                    st.stop()

                # Concatenate into one DataFrame
                df = pd.concat(all_dfs, ignore_index=True)

                # Persist raw uploads
                data_dir = os.path.join(BASE_DIR, 'data')
                os.makedirs(data_dir, exist_ok=True)
                for name, bytes_data in file_data:
                    dest = os.path.join(data_dir, name)
                    with open(dest, 'wb') as f:
                        f.write(bytes_data)

                # Update active config to first file
                first_dest = os.path.join(data_dir, file_data[0][0])
                with open(DATA_CONFIG, 'w') as f:
                    json.dump({'active_csv': first_dest}, f)

                # Generate & display supervisor codes
                code_map = load_or_create_codes(df)
                st.success("All CSVs uploaded and codes initialized.")
                st.subheader("Supervisor Access Codes")
                codes_df = (
                    pd.DataFrame.from_dict(code_map, orient='index', columns=['Code'])
                      .reset_index()
                      .rename(columns={'index':'Supervisor'})
                )
                st.table(codes_df)

                # Download codes as CSV
                csv_data = codes_df.to_csv(index=False)
                st.download_button(
                    "Download Supervisor Codes CSV", data=csv_data,
                    file_name="supervisor_codes.csv", mime="text/csv",
                    key="download_codes"
                )

    # Step 3: Download review log
        st.subheader("Download Access Review Log")
        if os.path.exists(EXCEL_OUTPUT):
            with open(EXCEL_OUTPUT, 'rb') as f:
                log_bytes = f.read()
            st.download_button(
                "Download Excel Log",
                data=BytesIO(log_bytes),
                file_name='access_review_log.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            st.info("No reviews logged yet.")

